from __future__ import annotations

import json
from io import BytesIO

import pytest
from openpyxl import load_workbook

from models import Country
from report import (
    _to_df,
    annex_a_payment_volumes,
    annex_b_fraud_breakdown,
    annex_c_sca_exemptions,
    annex_d_loss_allocation,
    build_report,
    render_json,
    render_parquet,
    render_xlsx,
)


def test_annex_a_aggregates_correctly(sample_rows) -> None:
    df = _to_df(sample_rows)
    section = annex_a_payment_volumes(df)
    assert section.annex == "A"
    # 2 distinct (instrument, channel) combos
    assert len(section.rows) == 2
    card_remote = next(r for r in section.rows if r["instrument"] == "card")
    assert card_remote["tx_count"] == 210_000
    assert card_remote["tx_value_eur"] == 85_000_000.0


def test_annex_b_filters_to_fraud_only(sample_rows) -> None:
    df = _to_df(sample_rows)
    section = annex_b_fraud_breakdown(df)
    assert all(r["fraud_count"] > 0 for r in section.rows)
    assert sum(r["fraud_count"] for r in section.rows) == 15


def test_annex_c_computes_fraud_rate_bps(sample_rows) -> None:
    df = _to_df(sample_rows)
    section = annex_c_sca_exemptions(df)
    tra_row = next(r for r in section.rows if r["sca_exemption"] == "tra")
    # 6500 / 5_000_000 * 10000 = 13.0 bps
    assert tra_row["fraud_rate_bps"] == pytest.approx(13.0)


def test_annex_d_groups_by_bearer_and_geo(sample_rows) -> None:
    df = _to_df(sample_rows)
    section = annex_d_loss_allocation(df)
    assert {r["loss_bearer"] for r in section.rows} == {"psp", "payer"}
    assert {r["counterparty_geo"] for r in section.rows} == {"eea", "non_eea"}


def test_build_report_totals(sample_rows) -> None:
    rep = build_report(Country.SE, "2025-Q1", sample_rows)
    assert rep.totals["tx_count"] == 260_000
    assert rep.totals["fraud_count"] == 15
    assert rep.totals["loss_value_eur"] == pytest.approx(4410.0)
    assert len(rep.sections) == 4
    assert rep.header.reporting_country == Country.SE
    assert rep.header.quarter == "2025-Q1"


def test_render_json_roundtrip(sample_rows) -> None:
    rep = build_report(Country.SE, "2025-Q1", sample_rows)
    payload = render_json(rep)
    parsed = json.loads(payload)
    assert parsed["header"]["eba_guideline"] == "EBA/GL/2020/01"
    assert parsed["header"]["schema_version"] == "1.2"
    assert len(parsed["sections"]) == 4


def test_render_xlsx_has_expected_sheets(sample_rows) -> None:
    rep = build_report(Country.SE, "2025-Q1", sample_rows)
    payload = render_xlsx(rep)
    wb = load_workbook(BytesIO(payload))
    assert set(wb.sheetnames) == {"Cover", "A", "B", "C", "D", "Totals"}
    cover = wb["Cover"]
    assert "EBA/GL/2020/01" in cover["A1"].value
    a = wb["A"]
    # Title + header row + at least one data row
    assert a["A1"].value.startswith("Annex A")
    assert a["A3"].value == "instrument"


def test_render_parquet_is_valid(sample_rows) -> None:
    import pyarrow.parquet as pq

    rep = build_report(Country.SE, "2025-Q1", sample_rows)
    payload = render_parquet(rep)
    table = pq.read_table(BytesIO(payload))
    cols = set(table.column_names)
    assert {
        "annex",
        "psp_lei",
        "reporting_country",
        "quarter",
        "submission_id",
    }.issubset(cols)


def test_empty_rows_produces_empty_sections() -> None:
    rep = build_report(Country.SE, "2025-Q1", [])
    for s in rep.sections:
        assert s.rows == []
    assert rep.totals["tx_count"] == 0
