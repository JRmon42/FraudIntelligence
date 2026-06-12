"""Generate the EBA/GL/2020/01 quarterly fraud report.

Pipeline:

    1. Read aggregated rows from the Fabric Lakehouse Delta table
       ``gold.eba_report_q`` (or, if unavailable, fall back to Cosmos DB
       ``analytics.eba_aggregates``).
    2. Build the four EBA Annex tables (A–D).
    3. Emit:
         * ``eba_<country>_<quarter>.xlsx``  (regulator-facing)
         * ``eba_<country>_<quarter>.json``  (machine-readable mirror)
         * ``eba_<country>_<quarter>.parquet`` (Power BI dataset shard)
       all uploaded to ADLS Gen2 under
       ``abfss://reports@{account}.dfs.core.windows.net/eba/{quarter}/{country}/``.
"""

from __future__ import annotations

import json
import logging
import os
import sys
import uuid
from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO
from pathlib import Path
from typing import Iterable

import click
import pandas as pd
import pyarrow as pa
import pyarrow.parquet as pq
import structlog
from azure.identity import DefaultAzureCredential
from azure.storage.filedatalake import DataLakeServiceClient
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pydantic_settings import BaseSettings, SettingsConfigDict
from tenacity import retry, stop_after_attempt, wait_exponential

from models import (
    AggregatedFraudRow,
    Country,
    EbaReport,
    Instrument,
    ReportHeader,
    ReportSection,
)

# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
structlog.configure(
    processors=[
        structlog.processors.add_log_level,
        structlog.processors.TimeStamper(fmt="iso"),
        structlog.processors.JSONRenderer(),
    ],
    wrapper_class=structlog.make_filtering_bound_logger(logging.INFO),
)
log = structlog.get_logger("eba-reporter")


# ---------------------------------------------------------------------------
# Settings
# ---------------------------------------------------------------------------
class Settings(BaseSettings):
    """Runtime configuration — sourced from env vars / KeyVault refs."""

    model_config = SettingsConfigDict(env_prefix="EBA_", case_sensitive=False)

    psp_lei: str = "5493001KJTIIGC8Y1R12"  # placeholder LEI
    psp_name: str = "Heimdall Nordic Payments AS"

    lakehouse_account: str = "fiprodlh"
    lakehouse_container: str = "lakehouse"
    lakehouse_path: str = "gold/eba_report_q"

    reports_account: str = "fiprodreports"
    reports_container: str = "reports"

    cosmos_endpoint: str = ""
    cosmos_database: str = "analytics"
    cosmos_container: str = "eba_aggregates"

    use_cosmos_fallback: bool = False


SETTINGS = Settings()

# ---------------------------------------------------------------------------
# Source readers
# ---------------------------------------------------------------------------


@retry(stop=stop_after_attempt(3), wait=wait_exponential(multiplier=1, min=2, max=10))
def read_from_lakehouse(country: Country, quarter: str) -> list[AggregatedFraudRow]:
    """Read aggregates from the Fabric Lakehouse Delta table."""
    from deltalake import DeltaTable  # local import — heavy

    cred = DefaultAzureCredential()
    storage_options = {
        "azure_storage_account_name": SETTINGS.lakehouse_account,
        "azure_storage_token": cred.get_token("https://storage.azure.com/.default").token,
    }
    uri = (
        f"abfss://{SETTINGS.lakehouse_container}@{SETTINGS.lakehouse_account}"
        f".dfs.core.windows.net/{SETTINGS.lakehouse_path}"
    )
    log.info("reading_lakehouse", uri=uri, country=country, quarter=quarter)
    dt = DeltaTable(uri, storage_options=storage_options)
    df = dt.to_pandas(
        partitions=[("reporting_country", "=", country.value), ("quarter", "=", quarter)]
    )
    return [AggregatedFraudRow.model_validate(r) for r in df.to_dict(orient="records")]


def read_from_cosmos(country: Country, quarter: str) -> list[AggregatedFraudRow]:
    """Fallback path: read aggregates from Cosmos DB (NoSQL API)."""
    from azure.cosmos import CosmosClient

    cred = DefaultAzureCredential()
    client = CosmosClient(SETTINGS.cosmos_endpoint, credential=cred)
    container = client.get_database_client(SETTINGS.cosmos_database).get_container_client(
        SETTINGS.cosmos_container
    )
    query = (
        "SELECT * FROM c WHERE c.reporting_country = @c AND c.quarter = @q"
    )
    items = container.query_items(
        query=query,
        parameters=[{"name": "@c", "value": country.value}, {"name": "@q", "value": quarter}],
        enable_cross_partition_query=True,
    )
    return [AggregatedFraudRow.model_validate(item) for item in items]


def load_rows(country: Country, quarter: str) -> list[AggregatedFraudRow]:
    if SETTINGS.use_cosmos_fallback:
        return read_from_cosmos(country, quarter)
    try:
        return read_from_lakehouse(country, quarter)
    except Exception as exc:  # pragma: no cover — defensive fallback
        log.warning("lakehouse_failed_falling_back_to_cosmos", error=str(exc))
        return read_from_cosmos(country, quarter)


# ---------------------------------------------------------------------------
# Aggregation — Annex builders
# ---------------------------------------------------------------------------


def _to_df(rows: Iterable[AggregatedFraudRow]) -> pd.DataFrame:
    df = pd.DataFrame([r.model_dump(mode="python") for r in rows])
    if df.empty:
        return df
    for col in ("tx_value_eur", "fraud_value_eur", "loss_value_eur"):
        df[col] = df[col].astype(float)
    return df


def annex_a_payment_volumes(df: pd.DataFrame) -> ReportSection:
    """Annex A — Total volumes & values by instrument & channel."""
    if df.empty:
        return ReportSection(annex="A", title="Annex A — Payment volumes & values", rows=[])
    grp = (
        df.groupby(["instrument", "channel"], as_index=False)
        .agg(tx_count=("tx_count", "sum"), tx_value_eur=("tx_value_eur", "sum"))
        .sort_values(["instrument", "channel"])
    )
    return ReportSection(
        annex="A",
        title="Annex A — Payment volumes & values",
        rows=grp.to_dict(orient="records"),
    )


def annex_b_fraud_breakdown(df: pd.DataFrame) -> ReportSection:
    """Annex B — Fraudulent transactions by instrument, channel, fraud type."""
    if df.empty:
        return ReportSection(annex="B", title="Annex B — Fraud breakdown", rows=[])
    fraud = df[df["fraud_count"] > 0]
    grp = (
        fraud.groupby(["instrument", "channel", "fraud_type"], as_index=False)
        .agg(
            fraud_count=("fraud_count", "sum"),
            fraud_value_eur=("fraud_value_eur", "sum"),
            loss_value_eur=("loss_value_eur", "sum"),
        )
        .sort_values(["instrument", "channel", "fraud_type"])
    )
    return ReportSection(
        annex="B",
        title="Annex B — Fraud breakdown",
        rows=grp.to_dict(orient="records"),
    )


def annex_c_sca_exemptions(df: pd.DataFrame) -> ReportSection:
    """Annex C — SCA application & exemption mix (incl. fraud rate per exemption)."""
    if df.empty:
        return ReportSection(annex="C", title="Annex C — SCA exemptions", rows=[])
    grp = (
        df.groupby(["instrument", "sca_exemption"], as_index=False)
        .agg(
            tx_count=("tx_count", "sum"),
            tx_value_eur=("tx_value_eur", "sum"),
            fraud_count=("fraud_count", "sum"),
            fraud_value_eur=("fraud_value_eur", "sum"),
        )
        .sort_values(["instrument", "sca_exemption"])
    )
    grp["fraud_rate_bps"] = (
        (grp["fraud_value_eur"] / grp["tx_value_eur"].replace(0, pd.NA)) * 10_000
    ).fillna(0).round(2)
    return ReportSection(
        annex="C",
        title="Annex C — SCA exemptions",
        rows=grp.to_dict(orient="records"),
    )


def annex_d_loss_allocation(df: pd.DataFrame) -> ReportSection:
    """Annex D — Loss allocation by bearer & geography."""
    if df.empty:
        return ReportSection(annex="D", title="Annex D — Loss allocation", rows=[])
    fraud = df[df["fraud_count"] > 0]
    grp = (
        fraud.groupby(["instrument", "loss_bearer", "counterparty_geo"], as_index=False)
        .agg(loss_value_eur=("loss_value_eur", "sum"), fraud_count=("fraud_count", "sum"))
        .sort_values(["instrument", "loss_bearer", "counterparty_geo"])
    )
    return ReportSection(
        annex="D",
        title="Annex D — Loss allocation",
        rows=grp.to_dict(orient="records"),
    )


def build_report(country: Country, quarter: str, rows: list[AggregatedFraudRow]) -> EbaReport:
    df = _to_df(rows)
    period_start, period_end = _quarter_to_dates(quarter)
    header = ReportHeader(
        psp_lei=SETTINGS.psp_lei,
        psp_name=SETTINGS.psp_name,
        reporting_country=country,
        quarter=quarter,
        period_start=period_start,
        period_end=period_end,
        submission_id=str(uuid.uuid4()),
    )
    sections = [
        annex_a_payment_volumes(df),
        annex_b_fraud_breakdown(df),
        annex_c_sca_exemptions(df),
        annex_d_loss_allocation(df),
    ]
    totals = {
        "tx_count": float(df["tx_count"].sum()) if not df.empty else 0.0,
        "tx_value_eur": float(df["tx_value_eur"].sum()) if not df.empty else 0.0,
        "fraud_count": float(df["fraud_count"].sum()) if not df.empty else 0.0,
        "fraud_value_eur": float(df["fraud_value_eur"].sum()) if not df.empty else 0.0,
        "loss_value_eur": float(df["loss_value_eur"].sum()) if not df.empty else 0.0,
    }
    return EbaReport(header=header, sections=sections, totals=totals)


def _quarter_to_dates(quarter: str) -> tuple[date, date]:
    year, q = quarter.split("-Q")
    y = int(year)
    q = int(q)
    starts = {1: (1, 1), 2: (4, 1), 3: (7, 1), 4: (10, 1)}
    ends = {1: (3, 31), 2: (6, 30), 3: (9, 30), 4: (12, 31)}
    return date(y, *starts[q]), date(y, *ends[q])


# ---------------------------------------------------------------------------
# XLSX rendering — programmatic openpyxl (no jinja2)
# ---------------------------------------------------------------------------

_HEADER_FILL = PatternFill("solid", fgColor="1F4E78")
_HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(bold=True, size=14, color="1F4E78")


def render_xlsx(report: EbaReport) -> bytes:
    wb = Workbook()
    # Sheet 1 — header / cover
    cover = wb.active
    cover.title = "Cover"
    cover["A1"] = "EBA/GL/2020/01 — Quarterly Fraud Report"
    cover["A1"].font = _TITLE_FONT
    meta = report.header.model_dump(mode="json")
    for i, (k, v) in enumerate(meta.items(), start=3):
        cover.cell(row=i, column=1, value=k).font = Font(bold=True)
        cover.cell(row=i, column=2, value=str(v))
    cover.column_dimensions["A"].width = 28
    cover.column_dimensions["B"].width = 60

    # Sheet per Annex
    for section in report.sections:
        ws = wb.create_sheet(title=section.annex)
        ws["A1"] = section.title
        ws["A1"].font = _TITLE_FONT
        ws.merge_cells("A1:H1")

        if not section.rows:
            ws["A3"] = "No data for this period."
            continue

        cols = list(section.rows[0].keys())
        for c, col in enumerate(cols, start=1):
            cell = ws.cell(row=3, column=c, value=col)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(c)].width = max(18, len(col) + 2)

        for r, row in enumerate(section.rows, start=4):
            for c, col in enumerate(cols, start=1):
                ws.cell(row=r, column=c, value=row[col])
        ws.freeze_panes = "A4"

    # Totals sheet
    tot = wb.create_sheet(title="Totals")
    tot["A1"] = "Period Totals"
    tot["A1"].font = _TITLE_FONT
    for i, (k, v) in enumerate(report.totals.items(), start=3):
        tot.cell(row=i, column=1, value=k).font = Font(bold=True)
        tot.cell(row=i, column=2, value=v)
    tot.column_dimensions["A"].width = 24
    tot.column_dimensions["B"].width = 24

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ---------------------------------------------------------------------------
# Parquet rendering — flattened wide table for Power BI
# ---------------------------------------------------------------------------


def render_parquet(report: EbaReport) -> bytes:
    rows = []
    for section in report.sections:
        for r in section.rows:
            rows.append({"annex": section.annex, **r})
    if not rows:
        rows = [{"annex": "A"}]
    df = pd.DataFrame(rows)
    df["psp_lei"] = report.header.psp_lei
    df["reporting_country"] = report.header.reporting_country.value
    df["quarter"] = report.header.quarter
    df["submission_id"] = report.header.submission_id
    df["generated_at_utc"] = datetime.now(timezone.utc).isoformat()
    table = pa.Table.from_pandas(df, preserve_index=False)
    buf = BytesIO()
    pq.write_table(table, buf, compression="snappy")
    return buf.getvalue()


def render_json(report: EbaReport) -> bytes:
    return report.model_dump_json(indent=2).encode("utf-8")


# ---------------------------------------------------------------------------
# ADLS Gen2 upload
# ---------------------------------------------------------------------------


def upload_to_adls(country: Country, quarter: str, name: str, payload: bytes) -> str:
    cred = DefaultAzureCredential()
    svc = DataLakeServiceClient(
        account_url=f"https://{SETTINGS.reports_account}.dfs.core.windows.net",
        credential=cred,
    )
    fs = svc.get_file_system_client(SETTINGS.reports_container)
    path = f"eba/{quarter}/{country.value}/{name}"
    file_client = fs.get_file_client(path)
    file_client.upload_data(payload, overwrite=True)
    log.info("uploaded", path=path, bytes=len(payload))
    return path


# ---------------------------------------------------------------------------
# CLI
# ---------------------------------------------------------------------------


@click.command()
@click.option(
    "--country",
    type=click.Choice([c.value for c in Country]),
    required=True,
    help="Reporting country (one of SE/NO/DK/FI/EE).",
)
@click.option("--quarter", required=True, help="Reporting quarter, e.g. 2025-Q1.")
@click.option(
    "--out",
    type=click.Path(file_okay=False, dir_okay=True, writable=True),
    default=None,
    help="Optional local output dir. If omitted, writes to ADLS only.",
)
@click.option("--dry-run", is_flag=True, help="Skip ADLS upload (local only).")
def cli(country: str, quarter: str, out: str | None, dry_run: bool) -> None:
    """Generate the EBA quarterly fraud report."""
    log.info("start", country=country, quarter=quarter, dry_run=dry_run)
    c = Country(country)
    rows = load_rows(c, quarter)
    log.info("rows_loaded", count=len(rows))

    report = build_report(c, quarter, rows)
    xlsx = render_xlsx(report)
    js = render_json(report)
    parquet = render_parquet(report)

    base = f"eba_{c.value}_{quarter}"
    artefacts = {
        f"{base}.xlsx": xlsx,
        f"{base}.json": js,
        f"{base}.parquet": parquet,
    }

    if out:
        outp = Path(out)
        outp.mkdir(parents=True, exist_ok=True)
        for n, payload in artefacts.items():
            (outp / n).write_bytes(payload)
            log.info("wrote_local", path=str(outp / n))

    if not dry_run:
        for n, payload in artefacts.items():
            upload_to_adls(c, quarter, n, payload)

    log.info("done", submission_id=report.header.submission_id)


if __name__ == "__main__":
    cli()
